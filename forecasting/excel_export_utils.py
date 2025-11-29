import logging
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional

# Константы (лучше вынести в отдельный файл constants.py и импортировать оттуда)
ID_SETTLEMENT_TOTAL = 1
ID_SETTLEMENT_URBAN = 2
ID_SETTLEMENT_RURAL = 3
SEX_CODE_TOTAL = "A"
SEX_CODE_MALE = "M"
SEX_CODE_FEMALE = "F"

logger = logging.getLogger(__name__)


def generate_forecast_excel_workbook(
        params_display_overall: Dict,
        all_warnings: List[str],
        grouped_forecasts_data: List[Dict],
        output_detailed_by_age_global: bool,
        user_selected_settlement_id: int,
        user_selected_sex_code: str
) -> Workbook:
    """
    Генерирует Excel книгу (Workbook) с данными прогноза.
    """
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # --- Стили ---
    bold_font_main_header = Font(bold=True, size=14)
    bold_font_subheader = Font(bold=True, size=12)
    param_label_font = Font(bold=True)
    center_aligned_text = Alignment(horizontal='center', vertical='center')
    wrap_text_alignment = Alignment(wrap_text=True, vertical='top')

    # --- ЛИСТ 1: Параметры и Предупреждения ---
    ws_params = wb.create_sheet(title="Параметры_и_Предупреждения")
    current_row = 1

    param_title_cell = ws_params.cell(row=current_row, column=1, value="Общие параметры исходного запроса")
    param_title_cell.font = bold_font_main_header
    ws_params.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    current_row += 2

    param_labels_map = {
        'region_names_display': "Изначально запрошенные регионы",
        'settlement_type_name_display': "Тип поселения (исходно)",
        'sex_code_name_display': "Целевой пол (исходно)",
        'forecast_period_display': "Период прогноза",
        'historical_period_display': "Период ист. данных",
        'target_age_group_input_display': "Возрастная группа (исходно)",
        'output_detailed_by_age_display': "Детализация по возрастам (на момент запроса)",
        'birth_rate_scenario_name_display': "Сценарий рождаемости",
        'death_rate_scenario_male_name_display': "Сценарий смертности (М)",
        'death_rate_scenario_female_name_display': "Сценарий смертности (Ж)",
        'include_migration_display': "Учет миграции",
        'migration_scenario_name_display': "Сценарий миграции",
    }

    display_values_for_params = params_display_overall.copy()
    if 'forecast_period_display' not in display_values_for_params:
        display_values_for_params[
            'forecast_period_display'] = f"{params_display_overall.get('forecast_start_year')} - {params_display_overall.get('forecast_end_year')}"
    if 'historical_period_display' not in display_values_for_params:
        display_values_for_params[
            'historical_period_display'] = f"{params_display_overall.get('historical_data_start_year')} - {params_display_overall.get('historical_data_end_year')}"
    if 'output_detailed_by_age_display' not in display_values_for_params:
        display_values_for_params['output_detailed_by_age_display'] = "Да" if output_detailed_by_age_global else "Нет"
    if 'include_migration_display' not in display_values_for_params:
        display_values_for_params['include_migration_display'] = "Да" if params_display_overall.get(
            'include_migration') else "Нет"

    for key, label in param_labels_map.items():
        value_to_write = display_values_for_params.get(key)
        if key == 'migration_scenario_name_display' and not display_values_for_params.get('include_migration'):
            continue
        if value_to_write is not None:
            if isinstance(value_to_write, list): value_to_write = ", ".join(map(str, value_to_write))
            ws_params.cell(row=current_row, column=1, value=label).font = param_label_font
            ws_params.cell(row=current_row, column=2, value=str(value_to_write))
            current_row += 1

    ws_params.column_dimensions[get_column_letter(1)].width = 45
    ws_params.column_dimensions[get_column_letter(2)].width = 65

    current_row += 1
    if all_warnings:
        warning_header_cell = ws_params.cell(row=current_row, column=1, value="Предупреждения и примечания")
        warning_header_cell.font = bold_font_subheader
        ws_params.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        current_row += 1
        for warning_idx, warning in enumerate(all_warnings):
            cell = ws_params.cell(row=current_row + warning_idx, column=1, value=warning)
            cell.alignment = wrap_text_alignment
            ws_params.merge_cells(start_row=current_row + warning_idx, start_column=1,
                                  end_row=current_row + warning_idx, end_column=3)  # Объединяем на 3 колонки
            ws_params.row_dimensions[current_row + warning_idx].height = (
                                                                                     len(warning) // 80 + 1) * 15  # Примерный автоподбор высоты
        current_row += len(all_warnings)

    # --- ЛИСТЫ ДЛЯ КАЖДОЙ ГРУППЫ РЕГИОНОВ ---
    for group_idx, group_data in enumerate(grouped_forecasts_data):
        clean_title = group_data['title'].replace('[', '').replace(']', '').replace('*', '').replace(':', '').replace(
            '?', '').replace('/', '-').replace('\\', '-')
        sheet_title_base = f"Прогноз_{clean_title}"[:25]  # Оставляем место для индекса, если что

        sheet_title = sheet_title_base
        title_suffix_idx = 1
        while sheet_title in wb.sheetnames:
            sheet_title = f"{sheet_title_base}_{title_suffix_idx}"[:30]  # Excel limit ~31 chars
            title_suffix_idx += 1
            if title_suffix_idx > 10:  # Fallback
                sheet_title = f"DataSheet_{group_idx}_{title_suffix_idx}"[:30]

        ws_data = wb.create_sheet(title=sheet_title)
        header_start_row = 1  # Строка для заголовка "Прогноз для..."
        data_header_row = header_start_row + 2  # Строка для "Городское/Сельское" и "Год/Возраст"
        sub_header_row = data_header_row + 1  # Строка для "Мужчины/Женщины/Всего"
        data_start_row = sub_header_row + 1  # Строка, с которой начинаются фактические данные

        # Заголовок группы регионов
        group_title_cell = ws_data.cell(row=header_start_row, column=1, value=f"Прогноз для: {group_data['title']}")
        group_title_cell.font = bold_font_main_header
        group_title_cell.alignment = center_aligned_text
        # Объединим этот заголовок позже, когда будем знать общую ширину таблицы

        # Формирование заголовков данных
        # Основные столбцы (Год, Возраст)
        main_data_cols = ['Год']
        if output_detailed_by_age_global:
            main_data_cols.append('Возраст')

        # Пишем основные заголовки "Год" и "Возраст"
        for i, col_title in enumerate(main_data_cols):
            cell = ws_data.cell(row=data_header_row, column=i + 1, value=col_title)
            cell.font = bold_font_subheader
            cell.alignment = center_aligned_text
            ws_data.merge_cells(start_row=data_header_row, start_column=i + 1,
                                end_row=sub_header_row, end_column=i + 1)  # Объединяем по вертикали

        current_col_for_data_headers = len(main_data_cols) + 1

        # Заголовки для "Городское население"
        if user_selected_settlement_id == ID_SETTLEMENT_TOTAL or user_selected_settlement_id == ID_SETTLEMENT_URBAN:
            colspan_urban = 0
            sub_headers_urban = []
            if user_selected_sex_code == SEX_CODE_TOTAL or user_selected_sex_code == SEX_CODE_MALE:
                sub_headers_urban.append('Мужчины');
                colspan_urban += 1
            if user_selected_sex_code == SEX_CODE_TOTAL or user_selected_sex_code == SEX_CODE_FEMALE:
                sub_headers_urban.append('Женщины');
                colspan_urban += 1
            if user_selected_sex_code == SEX_CODE_TOTAL:
                sub_headers_urban.append('Всего');
                colspan_urban += 1

            if colspan_urban > 0:
                urban_header_cell = ws_data.cell(row=data_header_row, column=current_col_for_data_headers)
                urban_header_cell.value = 'Городское население'
                urban_header_cell.font = bold_font_subheader
                urban_header_cell.alignment = center_aligned_text
                if colspan_urban > 1:
                    ws_data.merge_cells(start_row=data_header_row, start_column=current_col_for_data_headers,
                                        end_row=data_header_row,
                                        end_column=current_col_for_data_headers + colspan_urban - 1)

                for i, sub_h_val in enumerate(sub_headers_urban):
                    sub_h_cell = ws_data.cell(row=sub_header_row, column=current_col_for_data_headers + i)
                    sub_h_cell.value = sub_h_val
                    sub_h_cell.font = bold_font_subheader
                    sub_h_cell.alignment = center_aligned_text

                current_col_for_data_headers += colspan_urban

        # Заголовки для "Сельское население"
        if user_selected_settlement_id == ID_SETTLEMENT_TOTAL or user_selected_settlement_id == ID_SETTLEMENT_RURAL:
            colspan_rural = 0
            sub_headers_rural = []
            if user_selected_sex_code == SEX_CODE_TOTAL or user_selected_sex_code == SEX_CODE_MALE:
                sub_headers_rural.append('Мужчины');
                colspan_rural += 1
            if user_selected_sex_code == SEX_CODE_TOTAL or user_selected_sex_code == SEX_CODE_FEMALE:
                sub_headers_rural.append('Женщины');
                colspan_rural += 1
            if user_selected_sex_code == SEX_CODE_TOTAL:
                sub_headers_rural.append('Всего');
                colspan_rural += 1

            if colspan_rural > 0:
                rural_header_cell = ws_data.cell(row=data_header_row, column=current_col_for_data_headers)
                rural_header_cell.value = 'Сельское население'
                rural_header_cell.font = bold_font_subheader
                rural_header_cell.alignment = center_aligned_text
                if colspan_rural > 1:
                    ws_data.merge_cells(start_row=data_header_row, start_column=current_col_for_data_headers,
                                        end_row=data_header_row,
                                        end_column=current_col_for_data_headers + colspan_rural - 1)

                for i, sub_h_val in enumerate(sub_headers_rural):
                    sub_h_cell = ws_data.cell(row=sub_header_row, column=current_col_for_data_headers + i)
                    sub_h_cell.value = sub_h_val
                    sub_h_cell.font = bold_font_subheader
                    sub_h_cell.alignment = center_aligned_text

                current_col_for_data_headers += colspan_rural

        # Объединение заголовка группы регионов на всю ширину таблицы
        total_data_cols = current_col_for_data_headers - 1
        if total_data_cols > 0:  # Проверка, что есть хотя бы одна колонка данных
            ws_data.merge_cells(start_row=header_start_row, start_column=1, end_row=header_start_row,
                                end_column=total_data_cols)

        # Заполнение данными
        current_data_write_row = data_start_row
        for year_item in group_data['data_by_year']:
            year_val = year_item['year']
            if output_detailed_by_age_global:
                age_rows_for_year = year_item.get('age_rows', [])
                if not age_rows_for_year:  # Если нет детализации по возрасту для этого года, пропускаем
                    # Можно добавить запись суммарных данных, если они есть на уровне year_item
                    pass
                for age_row_idx, age_row_data in enumerate(age_rows_for_year):
                    col_idx = 1
                    # Год (только для первой возрастной строки)
                    ws_data.cell(row=current_data_write_row, column=col_idx, value=year_val if age_row_idx == 0 else "")
                    col_idx += 1
                    # Возраст
                    ws_data.cell(row=current_data_write_row, column=col_idx, value=age_row_data.get('age_display', '-'))
                    col_idx += 1

                    # Городские данные
                    if user_selected_settlement_id == ID_SETTLEMENT_TOTAL or user_selected_settlement_id == ID_SETTLEMENT_URBAN:
                        if user_selected_sex_code == SEX_CODE_TOTAL or user_selected_sex_code == SEX_CODE_MALE:
                            ws_data.cell(row=current_data_write_row, column=col_idx,
                                         value=age_row_data.get('urban_male', '-'));
                            col_idx += 1
                        if user_selected_sex_code == SEX_CODE_TOTAL or user_selected_sex_code == SEX_CODE_FEMALE:
                            ws_data.cell(row=current_data_write_row, column=col_idx,
                                         value=age_row_data.get('urban_female', '-'));
                            col_idx += 1
                        if user_selected_sex_code == SEX_CODE_TOTAL:
                            ws_data.cell(row=current_data_write_row, column=col_idx,
                                         value=age_row_data.get('urban_total', '-'));
                            col_idx += 1
                    # Сельские данные
                    if user_selected_settlement_id == ID_SETTLEMENT_TOTAL or user_selected_settlement_id == ID_SETTLEMENT_RURAL:
                        if user_selected_sex_code == SEX_CODE_TOTAL or user_selected_sex_code == SEX_CODE_MALE:
                            ws_data.cell(row=current_data_write_row, column=col_idx,
                                         value=age_row_data.get('rural_male', '-'));
                            col_idx += 1
                        if user_selected_sex_code == SEX_CODE_TOTAL or user_selected_sex_code == SEX_CODE_FEMALE:
                            ws_data.cell(row=current_data_write_row, column=col_idx,
                                         value=age_row_data.get('rural_female', '-'));
                            col_idx += 1
                        if user_selected_sex_code == SEX_CODE_TOTAL:
                            ws_data.cell(row=current_data_write_row, column=col_idx,
                                         value=age_row_data.get('rural_total', '-'));
                            col_idx += 1
                    current_data_write_row += 1
            else:  # Недетализированный по возрастам
                col_idx = 1
                ws_data.cell(row=current_data_write_row, column=col_idx, value=year_val);
                col_idx += 1
                # Городские данные
                if user_selected_settlement_id == ID_SETTLEMENT_TOTAL or user_selected_settlement_id == ID_SETTLEMENT_URBAN:
                    if user_selected_sex_code == SEX_CODE_TOTAL or user_selected_sex_code == SEX_CODE_MALE:
                        ws_data.cell(row=current_data_write_row, column=col_idx,
                                     value=year_item.get('urban_male', '-'));
                        col_idx += 1
                    if user_selected_sex_code == SEX_CODE_TOTAL or user_selected_sex_code == SEX_CODE_FEMALE:
                        ws_data.cell(row=current_data_write_row, column=col_idx,
                                     value=year_item.get('urban_female', '-'));
                        col_idx += 1
                    if user_selected_sex_code == SEX_CODE_TOTAL:
                        ws_data.cell(row=current_data_write_row, column=col_idx,
                                     value=year_item.get('urban_total', '-'));
                        col_idx += 1
                # Сельские данные
                if user_selected_settlement_id == ID_SETTLEMENT_TOTAL or user_selected_settlement_id == ID_SETTLEMENT_RURAL:
                    if user_selected_sex_code == SEX_CODE_TOTAL or user_selected_sex_code == SEX_CODE_MALE:
                        ws_data.cell(row=current_data_write_row, column=col_idx,
                                     value=year_item.get('rural_male', '-'));
                        col_idx += 1
                    if user_selected_sex_code == SEX_CODE_TOTAL or user_selected_sex_code == SEX_CODE_FEMALE:
                        ws_data.cell(row=current_data_write_row, column=col_idx,
                                     value=year_item.get('rural_female', '-'));
                        col_idx += 1
                    if user_selected_sex_code == SEX_CODE_TOTAL:
                        ws_data.cell(row=current_data_write_row, column=col_idx,
                                     value=year_item.get('rural_total', '-'));
                        col_idx += 1
                current_data_write_row += 1

        # Автоподбор ширины колонок для текущего листа данных
        for col_idx_dim in range(1, total_data_cols + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx_dim)
            for cell_dim in ws_data[col_letter]:  # Проходим по всем ячейкам колонки
                if cell_dim.value:
                    try:
                        cell_len = len(str(cell_dim.value))
                        if cell_len > max_len:
                            max_len = cell_len
                    except:  # pragma: no cover (на случай экзотических типов данных)
                        pass
            adjusted_width = max(max_len + 2, 10) if max_len > 0 else 10  # мин. ширина 10
            # Для объединенных ячеек заголовочных строк, max_len может быть некорректным.
            # Можно добавить проверку, что если это заголовочная ячейка, то брать ширину из значения.
            header_cell_value_data_row = ws_data.cell(row=data_header_row, column=col_idx_dim).value
            header_cell_value_sub_row = ws_data.cell(row=sub_header_row, column=col_idx_dim).value

            if header_cell_value_data_row and len(str(header_cell_value_data_row)) + 2 > adjusted_width:
                adjusted_width = len(str(header_cell_value_data_row)) + 4  # +4 для центрированных заголовков
            if header_cell_value_sub_row and len(str(header_cell_value_sub_row)) + 2 > adjusted_width:
                adjusted_width = len(str(header_cell_value_sub_row)) + 4

            ws_data.column_dimensions[col_letter].width = min(adjusted_width, 50)  # Ограничим макс. ширину

    return wb